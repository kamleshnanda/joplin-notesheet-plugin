#!/usr/bin/env python3
"""
Generate notesheet sample Excel files (#3-#13) for import/export testing.
Run: pip install openpyxl && python generate_samples.py
"""
import os
import random
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, CellIsRule, Rule, IconSet, FormatObject
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.formula import ArrayFormula

OUTPUT_DIR = "/Users/kamleshn/Desktop/notesheet-samples/"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def save(wb, name):
    path = os.path.join(OUTPUT_DIR, name)
    wb.save(path)
    size = os.path.getsize(path)
    print(f"  ✓ {name} ({size:,} bytes)")


# ─────────────────────────────────────────────────────────────
# 3. Hyperlinks-Variants.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Hyperlinks"
ws["A1"] = "Hyperlink Variants"
ws["A1"].font = Font(bold=True)

ws["A2"] = "https://example.com"
ws["A2"].hyperlink = "https://example.com"
ws["A2"].font = Font(color="0000FF", underline="single")

ws["A3"] = "mailto:hello@example.com"
ws["A3"].hyperlink = "mailto:hello@example.com"
ws["A3"].font = Font(color="0000FF", underline="single")

ws["A4"] = "file:///Users/kamleshn/Documents/sample.txt"
ws["A4"].hyperlink = "file:///Users/kamleshn/Documents/sample.txt"
ws["A4"].font = Font(color="0000FF", underline="single")

special_url = "https://x.com/a&b<c>d?e=1"
ws["A5"] = special_url
ws["A5"].hyperlink = special_url
ws["A5"].font = Font(color="0000FF", underline="single")

ws["A6"] = "Click here to visit our very long path page"
ws["A6"].hyperlink = "https://example.com/longpath/to/some/resource/that/is/quite/lengthy"
ws["A6"].font = Font(color="0000FF", underline="single")

dedup_url = "https://example.com/shared"
for col in ["A", "B", "C"]:
    cell = ws[f"{col}7"]
    cell.value = dedup_url
    cell.hyperlink = dedup_url
    cell.font = Font(color="0000FF", underline="single")

ws.column_dimensions["A"].width = 55
save(wb, "Hyperlinks-Variants.xlsx")


# ─────────────────────────────────────────────────────────────
# 4. BordersAndCellColors.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Borders and Colors"

border_styles = [
    "thin", "hair", "dotted", "dashed", "dashDot", "dashDotDot",
    "double", "medium", "mediumDashed", "mediumDashDot", "thick",
]

ws["A1"] = "Border Style"
ws["A1"].font = Font(bold=True)
ws["B1"] = "Sample"
ws["B1"].font = Font(bold=True)

for i, style in enumerate(border_styles, start=2):
    ws.cell(row=i, column=1).value = style
    cell = ws.cell(row=i, column=2)
    cell.value = f"Border: {style}"
    side = Side(style=style, color="000000")
    cell.border = Border(left=side, right=side, top=side, bottom=side)

row = len(border_styles) + 3

ws.cell(row=row, column=1).value = "Theme Font"
ws.cell(row=row, column=2).value = "Accent 4, lighter 40%"
ws.cell(row=row, column=2).font = Font(color=Color(theme=7, tint=0.39997558519241921))

row += 1
ws.cell(row=row, column=1).value = "Theme Border"
ws.cell(row=row, column=2).value = "Accent 1, tint"
ts = Side(style="medium", color=Color(theme=4, tint=0.39997558519241921))
ws.cell(row=row, column=2).border = Border(left=ts, right=ts, top=ts, bottom=ts)

row += 1
ws.cell(row=row, column=1).value = "RGB Font"
ws.cell(row=row, column=2).value = "Custom RGB #FF6600"
ws.cell(row=row, column=2).font = Font(color="FF6600")

row += 1
ws.cell(row=row, column=1).value = "RGB Fill"
ws.cell(row=row, column=2).value = "Custom RGB Fill #CCFFCC"
ws.cell(row=row, column=2).fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 30
save(wb, "BordersAndCellColors.xlsx")


# ─────────────────────────────────────────────────────────────
# 5. MergedCellsAndAlignment.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Merged and Alignment"

ws.merge_cells("A1:B2")
ws["A1"] = "Header"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("C1:D1")
ws["C1"] = "Center Middle"
ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

ws["A4"] = "This is a long text that should wrap within the cell boundaries to test wrap text behavior in Excel"
ws["A4"].alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 20

ws["B4"] = "Another wrapped\ntext with\nnewlines"
ws["B4"].alignment = Alignment(wrap_text=True)
ws.column_dimensions["B"].width = 20

ws["C4"] = "Short wrap"
ws["C4"].alignment = Alignment(wrap_text=True)

ws["A6"] = "Rotated 45 degrees"
ws["A6"].alignment = Alignment(text_rotation=45)

ws["B6"] = "Rotated 90 degrees"
ws["B6"].alignment = Alignment(text_rotation=90)

ws["C6"] = "Rotated -45 degrees"
ws["C6"].alignment = Alignment(text_rotation=135)
save(wb, "MergedCellsAndAlignment.xlsx")


# ─────────────────────────────────────────────────────────────
# 6. RichTextInOneCell.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "RichText"

ws["A1"] = CellRichText(TextBlock(InlineFont(b=True), "Hello"), " world")
ws["A2"] = CellRichText(
    TextBlock(InlineFont(color="FF0000"), "Red"),
    " and ",
    TextBlock(InlineFont(color="0000FF"), "Blue"),
    " text",
)
ws["A3"] = "Visit example.com for more info"
ws["A3"].hyperlink = "https://example.com"
ws["A3"].font = Font(color="0000FF", underline="single")
ws.column_dimensions["A"].width = 40
save(wb, "RichTextInOneCell.xlsx")


# ─────────────────────────────────────────────────────────────
# 7. NumberFormats.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Number Formats"

ws["A1"] = "Format Name"
ws["A1"].font = Font(bold=True)
ws["B1"] = "Value"
ws["B1"].font = Font(bold=True)
ws["C1"] = "Format Code"
ws["C1"].font = Font(bold=True)

formats = [
    ("General", 1234.5678, "General"),
    ("0", 1234.5678, "0"),
    ("0.00", 1234.5678, "0.00"),
    ("#,##0", 1234.5678, "#,##0"),
    ("#,##0.00", 1234.5678, "#,##0.00"),
    ("Currency USD", 1234.56, "$#,##0.00"),
    ("Currency EUR", 1234.56, '#,##0.00 "\u20ac"'),
    ("Currency SEK", 1234.56, '_-* #,##0.00 "kr"_-'),
    ("Percent 0%", 0.35, "0%"),
    ("Percent 0.00%", 0.3567, "0.00%"),
    ("Date m/d/yy", 45000, "m/d/yy"),
    ("Date yyyy-mm-dd", 45000, "yyyy-mm-dd"),
    ("Date dd-mmm-yy", 45000, "dd-mmm-yy"),
    ("Date AM/PM", 45000.75, '[$-409]m/d/yy h:mm AM/PM;@'),
    ("Accounting", 1234.56, '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'),
    ("Custom text @", "Hello", '@ "suffix"'),
    ("Conditional [Red]", -1234.56, "[Red]#,##0.00;[Blue]#,##0.00"),
]

for i, (label, value, fmt) in enumerate(formats, start=2):
    ws.cell(row=i, column=1).value = label
    cell = ws.cell(row=i, column=2)
    cell.value = value
    cell.number_format = fmt
    ws.cell(row=i, column=3).value = fmt

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 50
save(wb, "NumberFormats.xlsx")


# ─────────────────────────────────────────────────────────────
# 8. ConditionalFormatting-Variants.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "CF Variants"

ws["A1"] = "Color Scale (3-stop)"
ws["A1"].font = Font(bold=True)
for i in range(2, 12):
    ws.cell(row=i, column=1).value = (i - 2) * 10
ws.conditional_formatting.add("A2:A11", ColorScaleRule(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B"))

ws["C1"] = "Data Bar"
ws["C1"].font = Font(bold=True)
for i in range(2, 12):
    ws.cell(row=i, column=3).value = (i - 2) * 10
ws.conditional_formatting.add("C2:C11", DataBarRule(
    start_type="min", end_type="max", color="638EC6"))

ws["E1"] = "Highlight (>50 = Red)"
ws["E1"].font = Font(bold=True)
for i in range(2, 12):
    ws.cell(row=i, column=5).value = (i - 2) * 10
ws.conditional_formatting.add("E2:E11", CellIsRule(
    operator="greaterThan", formula=["50"],
    fill=PatternFill(bgColor="FFC7CE")))

ws["G1"] = "Top 3 (Green)"
ws["G1"].font = Font(bold=True)
random.seed(42)
for i in range(2, 12):
    ws.cell(row=i, column=7).value = random.randint(1, 100)
ws.conditional_formatting.add("G2:G11", Rule(
    type="top10", rank=3, percent=False, bottom=False,
    dxf=DifferentialStyle(fill=PatternFill(bgColor="C6EFCE"))))

ws["I1"] = "Icon Set (3 Arrows)"
ws["I1"].font = Font(bold=True)
for i in range(2, 12):
    ws.cell(row=i, column=9).value = (i - 2) * 10
ws.conditional_formatting.add("I2:I11", Rule(
    type="iconSet",
    iconSet=IconSet(
        iconSet="3Arrows",
        cfvo=[
            FormatObject(type="percent", val=0),
            FormatObject(type="percent", val=33),
            FormatObject(type="percent", val=67),
        ]
    )
))
save(wb, "ConditionalFormatting-Variants.xlsx")


# ─────────────────────────────────────────────────────────────
# 9. MultiSheet.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws_data = wb.active
ws_data.title = "Data"
ws_data["A1"] = "Category"
ws_data["B1"] = "Value"
ws_data["A1"].font = Font(bold=True)
ws_data["B1"].font = Font(bold=True)
for i, (c, v) in enumerate(
    [("Apples", 30), ("Bananas", 50), ("Cherries", 20),
     ("Dates", 45), ("Elderberry", 60)], 2):
    ws_data.cell(row=i, column=1).value = c
    ws_data.cell(row=i, column=2).value = v

ws_chart = wb.create_sheet("Chart")
ws_chart["A1"] = "Chart Sheet"
ws_chart["B2"] = 100
chart = BarChart()
chart.title = "Data Chart"
chart.style = 10
chart.add_data(Reference(ws_data, min_col=2, min_row=1, max_row=6), titles_from_data=True)
chart.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=6))
ws_chart.add_chart(chart, "A3")

ws_sum = wb.create_sheet("Summary")
ws_sum["A1"] = "Cross-sheet formula"
ws_sum["A1"].font = Font(bold=True)
ws_sum["A2"] = "=Data!B2+Chart!B2"
ws_sum["A3"] = "=SUM(Data!B2:B6)"
ws_sum["B1"] = "Description"
ws_sum["B1"].font = Font(bold=True)
ws_sum["B2"] = "Data!B2 + Chart!B2"
ws_sum["B3"] = "SUM of all Data values"
save(wb, "MultiSheet.xlsx")


# ─────────────────────────────────────────────────────────────
# 10. FormulasAndStructuredRefs.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Formulas"
ws["A1"] = "Col1"
ws["B1"] = "Col2"
ws["C1"] = "Col3"
for i in range(2, 12):
    ws.cell(row=i, column=1).value = i * 10
    ws.cell(row=i, column=2).value = i * 5
    ws.cell(row=i, column=3).value = f"Row{i}"

tab = Table(displayName="Table1", ref="A1:C11")
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

ws["E1"] = "Formula Type"
ws["E1"].font = Font(bold=True)
ws["F1"] = "Formula"
ws["F1"].font = Font(bold=True)

for i, (label, formula) in enumerate([
    ("Plain formula", "=A2+B2"),
    ("Range formula", "=SUM(A2:A11)"),
    ("Structured ref [Col1]", "=SUM(Table1[Col1])"),
    ("Structured ref #This Row", "=Table1[[#This Row],[Col1]]"),
    ("Structured ref #All", "=SUM(Table1[[#All],[Col2]])"),
], start=2):
    ws.cell(row=i, column=5).value = label
    ws.cell(row=i, column=6).value = formula

# Row 7: Array formula (CSE) spanning F7:F11
ws["E7"] = "Array formula (CSE)"
ws["F7"] = ArrayFormula("F7:F11", "=A2:A6*2")

# --- Create Sheet2 with Table2 (required for the cross-sheet ref) ---
ws2 = wb.create_sheet("Sheet2")
ws2["A1"] = "ColX"
ws2["B1"] = "ColY"
for i in range(2, 6):
    ws2.cell(row=i, column=1).value = i * 100
    ws2.cell(row=i, column=2).value = i * 50
tab2 = Table(displayName="Table2", ref="A1:B5")
tab2.tableStyleInfo = tab.tableStyleInfo
ws2.add_table(tab2)

# Row 13: Cross-sheet table ref (after the array spill range ends at F11)
ws["E13"] = "Cross-sheet table ref"
ws["F13"] = "=SUM(Table2[ColX])"

ws.column_dimensions["E"].width = 30
ws.column_dimensions["F"].width = 40
save(wb, "FormulasAndStructuredRefs.xlsx")


# ─────────────────────────────────────────────────────────────
# 11. LargeWorkbook.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Data"
headers = [f"Col_{i+1}" for i in range(20)]
for col_idx, h in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx).value = h
    ws.cell(row=1, column=col_idx).font = Font(bold=True)

random.seed(123)
for row in range(2, 1002):
    for col in range(1, 21):
        if col <= 5:
            ws.cell(row=row, column=col).value = round(random.uniform(1, 10000), 2)
        elif col <= 10:
            ws.cell(row=row, column=col).value = random.randint(1, 1000)
        elif col <= 15:
            ws.cell(row=row, column=col).value = "".join(
                random.choices(string.ascii_letters, k=10))
        else:
            ws.cell(row=row, column=col).value = random.choice(
                [True, False, None, round(random.uniform(-100, 100), 2)])

tab = Table(displayName="LargeTable", ref="A1:T1001")
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

ws.cell(row=1, column=21).value = "Formulas"
ws.cell(row=1, column=21).font = Font(bold=True)
for i in range(2, 102):
    if i % 4 == 0:
        ws.cell(row=i, column=21).value = f"=SUM(A{i}:E{i})"
    elif i % 4 == 1:
        ws.cell(row=i, column=21).value = f"=AVERAGE(F{i}:J{i})"
    elif i % 4 == 2:
        ws.cell(row=i, column=21).value = f"=A{i}*B{i}+C{i}"
    else:
        ws.cell(row=i, column=21).value = f'=IF(F{i}>500,"High","Low")'

chart = BarChart()
chart.title = "Performance Test Chart"
chart.style = 10
chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=20, max_col=3),
               titles_from_data=True)
chart.width = 20
chart.height = 12
ws_chart = wb.create_sheet("Chart")
ws_chart.add_chart(chart, "A1")
save(wb, "LargeWorkbook.xlsx")


# ─────────────────────────────────────────────────────────────
# 12. MaliciousValues.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Malicious"
ws["A1"] = "Test Case"
ws["A1"].font = Font(bold=True)
ws["B1"] = "Value"
ws["B1"].font = Font(bold=True)
ws["C1"] = "Notes"
ws["C1"].font = Font(bold=True)

ws["A2"] = "DDE Injection"
b2 = ws["B2"]
b2.value = "=cmd|'/c calc'!A1"
b2.data_type = "s"
ws["C2"] = "Must be preserved as text"

ws["A3"] = "Starts with ="
b3 = ws["B3"]
b3.value = "=1+1"
b3.data_type = "s"
ws["C3"] = "Formula-like string"

ws["A4"] = "Starts with +"
ws["B4"] = "+cmd|'/c calc'!A1"
ws["C4"] = "Plus prefix"

ws["A5"] = "Starts with -"
ws["B5"] = "-1+cmd|'/c calc'!A1"
ws["C5"] = "Minus prefix"

ws["A6"] = "Starts with @"
ws["B6"] = "@SUM(1+1)*cmd|'/c calc'!A1"
ws["C6"] = "At-sign prefix"

ws["A7"] = "NULL byte"
ws["B7"] = "before\\x00after"
ws["C7"] = "Placeholder - openpyxl rejects NULL. Patch binary or use xlsxwriter."

ws["A8"] = "Control chars"
ws["B8"] = "TAB[\t]LF[\n]CR[\r]"
ws["C8"] = "XML-legal control chars only (0x09, 0x0A, 0x0D)"

ws["A9"] = "1MB string"
ws["B9"] = "A" * (1024 * 1024)
ws["C9"] = "1,048,576 chars"

ws["A10"] = "Unicode RTL + emoji"
ws["B10"] = "\u202eevil\u202c \U0001f389 \ufeff BOM"
ws["C10"] = "RTL override, emoji, BOM"

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 60
ws.column_dimensions["C"].width = 60
save(wb, "MaliciousValues.xlsx")


# ─────────────────────────────────────────────────────────────
# 13. EmptyAndDegenerate.xlsx
# ─────────────────────────────────────────────────────────────
wb = Workbook()
ws1 = wb.active
ws1.title = "EmptySheet"
ws2 = wb.create_sheet("BareSheet")
ws3 = wb.create_sheet("HeaderOnlyTable")
ws3["A1"] = "Name"
ws3["B1"] = "Age"
ws3["C1"] = "City"
# Add one blank row so the table range is valid
ws3["A2"] = None
ws3["B2"] = None
ws3["C2"] = None
tab = Table(displayName="EmptyTable", ref="A1:C2")
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
ws3.add_table(tab)
save(wb, "EmptyAndDegenerate.xlsx")


print("\n✅ All 11 files generated successfully!")
